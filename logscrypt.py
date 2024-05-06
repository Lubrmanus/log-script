import requests
import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore
import re

def load_terms_and_patterns(excel_path, terms_sheet, patterns_sheet):
    """ Načte hledané termíny a regulární výrazy z Excel souboru. """
    wb = load_workbook(excel_path)
    terms_ws = wb[terms_sheet]
    patterns_ws = wb[patterns_sheet]
    
    search_terms = [cell.value for row in terms_ws.iter_rows(min_row=2, values_only=True) for cell in row if cell.value]
    search_patterns = [cell.value for row in patterns_ws.iter_rows(min_row=2, values_only=True) for cell in row if cell.value]
    
    wb.close()
    return search_terms, search_patterns

def search_logs(log_path, search_terms, search_patterns):
    """ Prohledá log soubor na základě seznamu hledaných termínů a regulárních výrazů. """
    with open(log_path, 'r', encoding='utf-8') as file:
        logs = file.readlines()
    
    results = {term: [] for term in search_terms}
    for term in search_terms:
        results[term] = [line for line in logs if term in line]
    
    for pattern in search_patterns:
        compiled_pattern = re.compile(pattern)
        results[pattern] = [line for line in logs if compiled_pattern.search(line)]
    
    return results

def send_to_opsgenie(alerts, api_key):
    url = "https://api.opsgenie.com/v2/alerts"
    headers = {
        "Authorization": "GenieKey " + api_key,
        "Content-Type": "application/json"
    }
    
    for alert, messages in alerts.items():
        if messages:  # Only send alert if there are messages
            message = f"Alert for {alert}: {len(messages)} occurrences found."
            post_data = {
                "message": message,
                "alias": alert,
                "description": "\n".join(messages),
                "priority": "P3"  # Set priority as needed
            }
            response = requests.post(url, headers=headers, json=post_data)
            print(f"Sent alert for {alert}. Status Code: {response.status_code}")

def main():
    # Cesta k Excel souboru a názvy listů
    excel_path = 'search_terms_patterns.xlsx'
    terms_sheet = 'Terms'
    patterns_sheet = 'Patterns'
    
    # Cesta k log souboru
    log_path = 'example.log'
    
    # Autorizačni klíč OG
    api_key = 'your_opsgenie_api_key_here'
    
    # Načtení hledaných termínů a regulárních výrazů
    search_terms, search_patterns = load_terms_and_patterns(excel_path, terms_sheet, patterns_sheet)
    
    # Prohledání log souboru
    search_results = search_logs(log_path, search_terms, search_patterns)
    
    # Odeslani do OG
    send_to_opsgenie(search_results, api_key)

if __name__ == "__main__":
    main()
